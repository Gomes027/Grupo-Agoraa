import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from extrair_lojas import Main
from config import DIR_CONTROLE_SMJ, DIR_CONTROLE_STT, DIR_CONTROLE_VIX, DIR_CONTROLE_MCP, DIR_TROCAS_SMJ, DIR_TROCAS_STT, DIR_TROCAS_VIX, DIR_TROCAS_MCP, ARQUIVO_EXCEL

class ManipuladorExcel:
    def __init__(self, nome_arquivo):
        self.nome_arquivo = nome_arquivo

    @staticmethod
    def ler_linha_especifica(arquivo):
        """Lê a última ou a antepenúltima linha de um arquivo Excel, dependendo do arquivo especificado."""
        df = pd.read_excel(arquivo, engine='xlrd')
        if len(df) > 0:
            if 'TROCAS - MCP' in arquivo:
                linha_especifica = df.iloc[[-3]]  # Antepenúltima linha para TROCAS_MCP
            else:
                linha_especifica = df.iloc[[-1]]  # Última linha para os demais arquivos
            colunas_nao_nulas = linha_especifica.dropna(axis=1)  # Remove colunas nulas da linha selecionada
            return colunas_nao_nulas
        else:
            return pd.DataFrame()  # Retorna um DataFrame vazio se o arquivo estiver vazio
        
    @staticmethod
    def renomear_colunas(df):
        novos_nomes_colunas = ['Registros', 'Soma Produto', 'Total Médio', 'Total Custo', 'Total Venda']
        df = df.iloc[:, 1:]  # Remove a primeira coluna
        df.columns = novos_nomes_colunas[:len(df.columns)]
        return df
    
    @staticmethod
    def adicionar_data_e_loja(df, arquivo):
        # Adiciona coluna 'Data' com a data atual
        df['Data'] = datetime.now().strftime('%d/%m/%Y')

        # Adiciona coluna 'Loja' baseada no nome do arquivo
        if 'SMJ' in arquivo:
            loja = 'SMJ'
        elif 'STT' in arquivo:
            loja = 'STT'
        elif 'VIX' in arquivo:
            loja = 'VIX'
        elif 'MCP' in arquivo:
            loja = 'MCP'
        else:
            loja = 'Desconhecida'
        df['Loja'] = loja

        # Reordena as colunas para que 'Data' e 'Loja' fiquem na frente
        colunas = ['Data', 'Loja'] + [coluna for coluna in df.columns if coluna not in ['Data', 'Loja']]
        df = df[colunas]

        return df
    
    @staticmethod
    def merge_e_separa_dfs(dataframes, identificadores):
        df_controle = pd.DataFrame()
        df_trocas = pd.DataFrame()
        
        # Itera sobre os DataFrames e seus respectivos identificadores
        for df, identificador in zip(dataframes, identificadores):
            if 'CONTROLE' in identificador:
                # Faz merge com o DataFrame de 'CONTROLE' se ainda não estiver vazio, caso contrário, inicializa com o atual
                df_controle = pd.concat([df_controle, df], ignore_index=True) if not df_controle.empty else df
            elif 'TROCAS' in identificador:
                # Faz merge com o DataFrame de 'TROCAS' se ainda não estiver vazio, caso contrário, inicializa com o atual
                df_trocas = pd.concat([df_trocas, df], ignore_index=True) if not df_trocas.empty else df
        
        return df_controle, df_trocas
    
class SalvarExcel:
    def __init__(self, nome_arquivo):
        self.nome_arquivo = nome_arquivo

    def salvar_ou_atualizar_excel(self, df_controle, df_trocas):
        if os.path.exists(self.nome_arquivo):
            self.atualizar_excel(df_controle, 'Lojas de Controle')
            self.atualizar_excel(df_trocas, 'Lojas de Trocas')
            self.ajustar_largura_colunas_openpyxl('Lojas de Controle')
            self.ajustar_largura_colunas_openpyxl('Lojas de Trocas')
        else:
            with pd.ExcelWriter(self.nome_arquivo, engine='xlsxwriter') as writer:
                df_controle.to_excel(writer, sheet_name='Lojas de Controle', index=False)
                SalvarExcel.ajustar_largura_colunas(writer, df_controle, 'Lojas de Controle')
                df_trocas.to_excel(writer, sheet_name='Lojas de Trocas', index=False)
                SalvarExcel.ajustar_largura_colunas(writer, df_trocas, 'Lojas de Trocas')

    @staticmethod
    def ajustar_largura_colunas(writer, df, sheet_name):
        if writer.engine == 'xlsxwriter':
            worksheet = writer.sheets[sheet_name]
            for i, col in enumerate(df.columns):
                col_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(i, i, col_width)

    def ajustar_largura_colunas_openpyxl(self, sheet_name):
        wb = load_workbook(self.nome_arquivo)
        ws = wb[sheet_name]

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        wb.save(self.nome_arquivo)

    def atualizar_excel(self, df_novo, sheet_name):
        if sheet_name in pd.ExcelFile(self.nome_arquivo).sheet_names:
            df_existente = pd.read_excel(self.nome_arquivo, sheet_name=sheet_name)
            # Concatena e remove duplicatas com base em 'Data' e 'Loja', ajuste conforme necessário
            df_atualizado = self.atualizar_df_com_condicao(df_existente, df_novo)
            with pd.ExcelWriter(self.nome_arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_atualizado.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.book.save(self.nome_arquivo)  # Salvar o livro Excel diretamente
        else:
            with pd.ExcelWriter(self.nome_arquivo, engine='openpyxl', mode='a') as writer:
                df_novo.to_excel(writer, sheet_name=sheet_name, index=False)
                writer.book.save(self.nome_arquivo)  # Salvar o livro Excel diretamente

    @staticmethod
    def atualizar_df_com_condicao(df_existente, df_novo):
        # Verifica se as datas e lojas são diferentes e então concatena, caso contrário substitui
        df_concatenado = pd.concat([df_existente, df_novo])
        df_atualizado = df_concatenado.drop_duplicates(subset=['Data', 'Loja'], keep='last')
        return df_atualizado
    
class ControladorDeFluxo:
    def __init__(self, arquivo_excel):
        self.arquivo_excel = arquivo_excel
        self.dfs = []  # Lista para coletar DataFrames processados
        self.identificadores = []  # Lista para coletar identificadores

    def processar_arquivos(self, arquivo_controle):
        """Processa cada arquivo de controle para preparar os DataFrames."""
        for arquivo in arquivo_controle:
            linha_especifica = ManipuladorExcel.ler_linha_especifica(arquivo)
            df_renomeado = ManipuladorExcel.renomear_colunas(linha_especifica)
            df_novas_colunas = ManipuladorExcel.adicionar_data_e_loja(df_renomeado, arquivo)
            
            # Adiciona o DataFrame processado e o identificador à suas respectivas listas
            self.dfs.append(df_novas_colunas)
            self.identificadores.append(arquivo)  # Ou use uma lógica para extrair "CONTROLE" ou "TROCAS" do nome do arquivo

            os.remove(arquivo)

    def executar(self):
        """Executa o processo completo de manipulação dos arquivos Excel."""
        manipulador = SalvarExcel(self.arquivo_excel)

        # Processa os arquivos de controle e trocas
        self.processar_arquivos([DIR_CONTROLE_SMJ, DIR_CONTROLE_STT, DIR_CONTROLE_VIX, DIR_CONTROLE_MCP, DIR_TROCAS_SMJ, DIR_TROCAS_STT, DIR_TROCAS_VIX, DIR_TROCAS_MCP])

        # Agora, chama merge_e_separa_dfs com as listas coletadas
        df_controle, df_trocas = ManipuladorExcel.merge_e_separa_dfs(self.dfs, self.identificadores)

        # Usa a classe SalvarExcel para salvar ou atualizar o Excel
        manipulador.salvar_ou_atualizar_excel(df_controle, df_trocas)

if __name__ == "__main__":
    controlador = ControladorDeFluxo(ARQUIVO_EXCEL)
    extrair_lojas = Main()
    extrair_lojas.executar()
    controlador.executar()