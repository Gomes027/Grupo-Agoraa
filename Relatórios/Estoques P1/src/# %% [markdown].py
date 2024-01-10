# %% [markdown]
# Importa o relatório tresmann e o relatório de preços.

# %%
import warnings
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter

# Ignorar apenas o aviso específico mencionado
warnings.filterwarnings("ignore", message="Workbook contains no default style, apply openpyxl's default")
warnings.filterwarnings('ignore', category=UserWarning, module='pandas.core.tools.datetimes')


# Importar Arquivos excel como Dataframes
df_relatorio = pd.read_excel(r"F:\COMPRAS\relatorio_tresmann.xlsx")
df_preços = pd.read_excel(r"F:\BI\Bases\relatorio_precos_4_lojas.xlsx")

# %% [markdown]
# Aplica filtros no relatório tresmann e remove algumas colunas do dataframe de preços.

# %%
def filtrar_relatorio(df):
    """
    Filtra o DataFrame do relatório com base sem critérios específicos:
    - Exclui código 83566
    - Exclui setores "MATERIAL CONSUMO" e "HORTIFRUTI"
    - Exclui loja "MERCAPP"
    - Mantém apenas itens que não estão "FORA DO MIX"
    - Exclui nomes de fantasia específicos
    """
    nomes_fantasia_excluidos = [
        "COCONUT LTDA", "DOMART ALIMENTOS LTDA", "FORTE BOI",
        "PRO LARANJA", "SUIMARTIN INDUSTRIA E COMERCIO", 
        "TRESMANN - SMJ", "TRESMANN - STT"
    ]

    df_filtrado = df[
        (df["CODIGO"] != 83566) &
        (df["SETOR"] != "MATERIAL CONSUMO") &
        (df["LOJA"] != "MERCAPP") &
        (df["FORA DO MIX"] == "NAO") &
        (df["SETOR"] != "HORTIFRUTI") &
        ~df["NOME_FANTASIA"].isin(nomes_fantasia_excluidos)
    ]

    return df_filtrado

def filtrar_precos(df):
    """
    Remove colunas específicas do DataFrame de preços.
    """
    colunas_a_remover = ["NOME", "LOJA", "PRECOPROMOCAO", "PRECOVENDA2"]
    return df.drop(colunas_a_remover, axis=1)

# Aplicando as funções de filtragem
df_relatorio_filtrado = filtrar_relatorio(df_relatorio)
df_preços_filtrado = filtrar_precos(df_preços)

# %% [markdown]
# Converte tipos de dados, trata valores ausentes do relatório tresmann e renomeia algumas colunas de ambos dataframes.

# %%
def converter_dados_relatorio(df):
    """
    Converte tipos de dados e trata valores ausentes no DataFrame do relatório.
    - Converte "CODG DE BARRAS", "PEDIDO EM ABERTO (QTDE)" e "EMB_TRANSF" para Int64
    - Trata valores ausentes em "PEDIDO EM ABERTO (QTDE)" e "FLAG5"
    - Converte "VALIDADE" para formato datetime
    - Renomeia coluna "Ã¿LTIMO CUSTO ("
    """
    df["CODG DE BARRAS"] = df["CODG DE BARRAS"].astype("Int64")
    df["PEDIDO EM ABERTO (QTDE)"] = df["PEDIDO EM ABERTO (QTDE)"].fillna(0).round().astype("Int64")
    df["VALIDADE"] = pd.to_datetime(df["VALIDADE"], errors="coerce")
    df["EMB_TRANSF"] = df["EMB_TRANSF"].astype("Int64")
    df["FLAG5"] = pd.to_numeric(df["FLAG5"], errors="coerce").fillna(0).astype("Int64")
    df = df.rename(columns={"Ã¿LTIMO CUSTO (": "ULTIMO CUSTO"})
    return df

def renomear_colunas_precos(df):
    """
    Renomeia colunas no DataFrame de preços para nomes mais legíveis.
    """
    colunas_renomeadas = {
        "NOME_LOJA": "LOJA",
        "CUSTOLIQUIDO": "PREÇO DE CUSTO",
        "PRECOVENDA": "PREÇO DE VENDA"
    }
    return df.rename(columns=colunas_renomeadas)

# Aplicando as funções de limpeza e renomeação
df_relatorio_renomeado = converter_dados_relatorio(df_relatorio_filtrado.copy())
df_preços_renomeado = renomear_colunas_precos(df_preços_filtrado)

# %% [markdown]
# Realiza operações de calculos, como vencimento, media, estoque e comparação.

# %%
def calcular_dias_para_vencer(df):
    """
    Calcula a diferença em dias entre a data de validade e a data atual.
    """
    AQ1 = datetime.now().date()
    df["VALIDADE"] = pd.to_datetime(df["VALIDADE"]).dt.date
    df["DIAS PARA VENCER"] = df["VALIDADE"].apply(lambda x: (x - AQ1).days if not pd.isna(x) and (x - AQ1).days > 0 else 0)
    return df

def calcular_medias_e_projeções(df):
    """
    Calcula vendas médias mensais, vendas diárias, projeção de vendas e verifica se o estoque atual é maior que a projeção.
    """
    df["VENDA MEDIA MENSAL"] = df["VENDA ULTIMOS 12 MESES (QTDE)"] / 12
    df["MAIOR VENDA"] = df[["VENDA ULTIMOS 30 DIAS (QTDE)", "VENDA MEDIA MENSAL"]].max(axis=1)
    df["VENDA DIÁRIA"] = df["MAIOR VENDA"] / 30
    df["PROJEÇÃO DE VENDA"] = df["DIAS PARA VENCER"] * df["VENDA DIÁRIA"]
    df["EST > PROJ"] = df.apply(lambda row: 1 if row["ESTOQUE ATUAL"] > row["PROJEÇÃO DE VENDA"] else 0, axis=1)
    return df

def calcular_estoque_excedente_e_custos(df):
    """
    Calcula o estoque excedente, custo total e venda total.
    """
    df["ESTOQUE EXCEDENTE"] = df.apply(lambda row: 0 if row["EST > PROJ"] == 0 else row["ESTOQUE ATUAL"] - row["PROJEÇÃO DE VENDA"], axis=1)
    df["CUSTO TOTAL"] = df["ESTOQUE EXCEDENTE"] * df["PREÇO DE CUSTO"]
    df["VENDA TOTAL"] = df["PREÇO DE VENDA"] * df["ESTOQUE EXCEDENTE"]
    return df

def adicionar_colunas_comparativas(df):
    """
    Adiciona colunas comparativas para a validade de 60 dias e quantidade em estoque versus embalagem.
    """
    df["60 DIAS < PRAZO"] = df["DIAS PARA VENCER"].apply(lambda x: "SIM" if x <= 60 else "NÃO")
    df["EST ATUAL < X EMB"] = df.apply(lambda row: "SIM" if row["ESTOQUE ATUAL"] < (1 * row["QTDE NA EMBALAGEM"]) else "NÃO", axis=1)
    return df

# Aplicando as funções ao DataFrame
df_relatorio_calculos = df_relatorio_renomeado.copy()
df_relatorio_calculos = calcular_dias_para_vencer(df_relatorio_calculos)
df_relatorio_calculos = calcular_medias_e_projeções(df_relatorio_calculos)
df_relatorio_calculos = df_relatorio_calculos.merge(df_preços_renomeado, on=["CODIGO", "LOJA"], how="left")
df_relatorio_calculos = calcular_estoque_excedente_e_custos(df_relatorio_calculos)
df_relatorio_calculos = adicionar_colunas_comparativas(df_relatorio_calculos)

df_relatorio_calculos

# %% [markdown]
# Faz um Dataframe com o resumo dos dados apresentados.

# %%
def agregar_dados_por_loja(df):
    """
    Agrupa o dataframe por "LOJA" e calcula as quantidades de itens, estoque excedente,
    custo total e venda total. Retorna um dataframe agregado.
    """
    agregados = {
        "EST > PROJ": lambda x: (x == 1).sum(),
        "ESTOQUE EXCEDENTE": "sum",  
        "CUSTO TOTAL": "sum",        
        "VENDA TOTAL": "sum"        
    }
    
    # Realizar o agrupamento e a agregação
    df_agregado = df.groupby("LOJA").agg(agregados).reset_index()
    
    # Renomear as colunas do DataFrame resultante para corresponder à saída desejada
    df_agregado.columns = ["LOJA", "QTDE ITENS", "QTDE EXCEDENTE", "VALOR CUSTO TOTAL", "VALOR VENDA TOTAL"]
    
    return df_agregado

def adicionar_linha_total(df):
    """
    Adiciona uma linha de totais ao final do dataframe.
    """
    total_itens = df["QTDE ITENS"].sum()
    total_excedente = df["QTDE EXCEDENTE"].sum()
    total_custo = df["VALOR CUSTO TOTAL"].sum()
    total_venda = df["VALOR VENDA TOTAL"].sum()
    df.loc["Total"] = ["", total_itens, total_excedente, total_custo, total_venda]
    return df

def formatar_qtde_excedente(df):
    """
    Converte a coluna "QTDE EXCEDENTE" para inteiro.
    """
    df["QTDE EXCEDENTE"] = pd.to_numeric(df["QTDE EXCEDENTE"], errors="coerce").fillna(0).astype(int)
    return df


# Aplicando as funções
df_relatorio_resumo = df_relatorio_calculos.copy()
df_relatorio_resumo = agregar_dados_por_loja(df_relatorio_resumo)
df_relatorio_resumo = adicionar_linha_total(df_relatorio_resumo)
df_relatorio_resumo = formatar_qtde_excedente(df_relatorio_resumo)

df_relatorio_resumo

# %% [markdown]
# Processa a ánalise diaria dos itens com Estoque Excedente por loja.

# %%
def processar_dados_por_loja(df, loja):
    # Garantir que "ESTOQUE EXCEDENTE" esteja no tipo correto e tratar dados faltantes
    df["ESTOQUE EXCEDENTE"] = pd.to_numeric(df["ESTOQUE EXCEDENTE"], errors="coerce").fillna(0).astype(int)

    # Filtrar por loja e por estoque excedente maior que zero e custo total maior que zero
    df_filtrado = df[((df["LOJA"] == loja) & (df["CUSTO TOTAL"] > 0)) & (df["ESTOQUE EXCEDENTE"] > 0)].copy()

    # Selecionar colunas específicas e ordenar
    df_ordenado = df_filtrado[["NOME", "CODIGO", "SETOR", "ESTOQUE EXCEDENTE", "CUSTO TOTAL"]].sort_values(by=["CUSTO TOTAL"], ascending=False)

    return df_ordenado

# Exemplo de uso:
resultado_loja_SMJ = processar_dados_por_loja(df_relatorio_calculos, "TRESMANN - SMJ")
resultado_loja_STT = processar_dados_por_loja(df_relatorio_calculos, "TRESMANN - STT")
resultado_loja_VIX = processar_dados_por_loja(df_relatorio_calculos, "TRESMANN - VIX")

# %% [markdown]
# Ajusta a largura das colunas e tipo de dados do arquivo Excel, depois salvas as tabelas.

# %%
# Constantes
CAMINHO_ARQUIVO = 'relatorio_lojas.xlsx'
COLUNAS_MOEDA = ["VALOR CUSTO TOTAL", "VALOR VENDA TOTAL", "CUSTO TOTAL"]
COLUNAS_SEPARADOR_MILHAR = ["QTDE ITENS", "QTDE EXCEDENTE", "ESTOQUE EXCEDENTE"]

def ajustar_formato_colunas(escritor, nome_aba, dataframe, colunas_moeda, colunas_milhar):
    """Ajusta a largura e o formato das colunas."""
    for coluna in dataframe:
        largura_coluna = max(dataframe[coluna].astype(str).map(len).max(), len(coluna)) + 2
        indice_coluna = dataframe.columns.get_loc(coluna) + 1
        escritor.sheets[nome_aba].column_dimensions[get_column_letter(indice_coluna)].width = largura_coluna

        if colunas_moeda and coluna in colunas_moeda:
            aplicar_formato_moeda(escritor, nome_aba, indice_coluna, dataframe.shape[0])

def aplicar_formato_moeda(escritor, nome_aba, indice_coluna, quantidade_linhas):
    """Aplica o formato de moeda às células."""
    for linha in range(2, quantidade_linhas + 2):
        celula = escritor.sheets[nome_aba].cell(row=linha, column=indice_coluna)
        celula.number_format = "R$#,##0.00"

def salvar_dfs_no_excel(escritor, informacoes_dfs):
    """Salva os DataFrames em suas respectivas abas e ajusta as colunas."""
    for info_df in informacoes_dfs:
        info_df['dataframe'].to_excel(escritor, sheet_name=info_df['nome_aba'], index=False)
        ajustar_formato_colunas(escritor, info_df['nome_aba'], info_df['dataframe'], COLUNAS_MOEDA, COLUNAS_SEPARADOR_MILHAR)

def anexar_aba_historico(df_novo_historico):
    """Anexa os novos dados à aba 'HISTÓRICO P1'."""
    try:
        df_historico_existente = pd.read_excel(CAMINHO_ARQUIVO, sheet_name='HISTÓRICO P1')
    except ValueError:  # Se a aba não existir, criar um DataFrame vazio
        df_historico_existente = pd.DataFrame()

    df_historico_combinado = pd.concat([df_historico_existente, df_novo_historico], ignore_index=True)

    # Formata a coluna 'DATA' no formato 'dd/mm/aaaa'
    df_historico_combinado['DATA'] = pd.to_datetime(df_historico_combinado['DATA'], format='%d/%m/%Y', errors='coerce').dt.strftime('%d/%m/%Y')
    
    with pd.ExcelWriter(CAMINHO_ARQUIVO, engine='openpyxl', mode='a', if_sheet_exists='replace') as escritor:
        df_historico_combinado.to_excel(escritor, sheet_name='HISTÓRICO P1', index=False)
        ajustar_formato_colunas(escritor, 'HISTÓRICO P1', df_historico_combinado, COLUNAS_MOEDA, COLUNAS_SEPARADOR_MILHAR)

def verificar_existencia_arquivo():
    """Verifica se o arquivo Excel existe."""
    try:
        # Tenta ler o arquivo
        pd.read_excel(CAMINHO_ARQUIVO, engine='openpyxl')
        return True
    except FileNotFoundError:
        # Se o arquivo não existir, retorna False
        return False

# Código principal
if verificar_existencia_arquivo():
    # Se o arquivo existir, ler a aba 'RESUMO' e preparar os dados para 'HISTÓRICO P1'
    df_resumo_relatorio = pd.read_excel(CAMINHO_ARQUIVO, engine='openpyxl', sheet_name='RESUMO')
    df_novo_historico = df_resumo_relatorio.iloc[:-1].copy()
    df_novo_historico['DATA'] = datetime.now().strftime('%d/%m/%Y')
    anexar_aba_historico(df_novo_historico)
else:
    # Se o arquivo não existir, criar um novo arquivo e salvar os DataFrames
    with pd.ExcelWriter(CAMINHO_ARQUIVO, engine='openpyxl') as writer:
        # Suponha que você tenha os seguintes DataFrames para salvar: df_resumo, df_loja_SMJ, df_loja_STT, df_loja_VIX
        # Substitua com seus próprios DataFrames
        df_relatorio_resumo.to_excel(writer, sheet_name='RESUMO', index=False)
        ajustar_formato_colunas(writer, 'RESUMO', df_relatorio_resumo, COLUNAS_MOEDA, COLUNAS_SEPARADOR_MILHAR)

        resultado_loja_SMJ.to_excel(writer, sheet_name='TRESMANN - SMJ', index=False)
        ajustar_formato_colunas(writer, 'TRESMANN - SMJ', resultado_loja_SMJ, COLUNAS_MOEDA, COLUNAS_SEPARADOR_MILHAR)

        resultado_loja_STT.to_excel(writer, sheet_name='TRESMANN - STT', index=False)
        ajustar_formato_colunas(writer, 'TRESMANN - STT', resultado_loja_STT, COLUNAS_MOEDA, COLUNAS_SEPARADOR_MILHAR)

        resultado_loja_VIX.to_excel(writer, sheet_name='TRESMANN - VIX', index=False)
        ajustar_formato_colunas(writer, 'TRESMANN - VIX', resultado_loja_VIX, COLUNAS_MOEDA, COLUNAS_SEPARADOR_MILHAR)