import warnings
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from enviar_email import enviar_email

# Ignorar avisos específicos
warnings.filterwarnings("ignore", message="Workbook contains no default style, apply openpyxl's default")
warnings.filterwarnings('ignore', category=UserWarning, module='pandas.core.tools.datetimes')

# Variáveis globais
colunas_moeda = ["VALOR CUSTO TOTAL", "VALOR VENDA TOTAL", "CUSTO TOTAL EXCEDENTE"]
CAMINHO_ARQUIVO = r'F:\BI\ESTOQUE_EXCEDENTE.xlsx'

# Importar Arquivos excel como Dataframes
df_relatorio = pd.read_excel(r"F:\COMPRAS\relatorio_tresmann.xlsx")
df_preços = pd.read_excel(r"F:\BI\Bases\relatorio_precos_4_lojas.xlsx")

# Funções de Filtragem
def filtrar_relatorio(df):
    nomes_fantasia_excluidos = [
        "COCONUT LTDA", "DOMART ALIMENTOS LTDA", "FORTE BOI",
        "PRO LARANJA", "SUIMARTIN INDUSTRIA E COMERCIO", 
        "TRESMANN - SMJ", "TRESMANN - STT"
    ]
    df_filtrado = df[
        (df["CODIGO"] != 83566) &
        (df["SETOR"] != "MATERIAL CONSUMO") &
        (df["LOJA"] != "MERCAPP") &
        (df["FORA DO MIX"] == "NAO") &
        (df["SETOR"] != "HORTIFRUTI") &
        ~df["NOME_FANTASIA"].isin(nomes_fantasia_excluidos)
    ]
    return df_filtrado

def filtrar_precos(df):
    colunas_a_remover = ["NOME", "LOJA", "PRECOPROMOCAO", "PRECOVENDA2"]
    return df.drop(colunas_a_remover, axis=1)

# Aplicando as funções de filtragem
df_relatorio_filtrado = filtrar_relatorio(df_relatorio)
df_preços_filtrado = filtrar_precos(df_preços)

# Funções de Conversão e Renomeação
def converter_dados_relatorio(df):
    df = df.copy()
    
    df.loc[:, "CODG DE BARRAS"] = df["CODG DE BARRAS"].astype("Int64")
    df.loc[:, "PEDIDO EM ABERTO (QTDE)"] = df["PEDIDO EM ABERTO (QTDE)"].fillna(0).round().astype("Int64")
    df.loc[:, "VALIDADE"] = pd.to_datetime(df["VALIDADE"], errors="coerce")
    df.loc[:, "EMB_TRANSF"] = df["EMB_TRANSF"].astype("Int64")
    df.loc[:, "FLAG5"] = pd.to_numeric(df["FLAG5"], errors="coerce").fillna(0).astype("Int64")
    df = df.rename(columns={"Ã¿LTIMO CUSTO (": "ULTIMO CUSTO"})
    return df

def renomear_colunas_precos(df):
    colunas_renomeadas = {
        "NOME_LOJA": "LOJA",
        "CUSTOLIQUIDO": "PREÇO DE CUSTO",
        "PRECOVENDA": "PREÇO DE VENDA"
    }
    return df.rename(columns=colunas_renomeadas)

# Aplicando as funções de conversão e renomeação
df_relatorio_renomeado = converter_dados_relatorio(df_relatorio_filtrado)
df_preços_renomeado = renomear_colunas_precos(df_preços_filtrado)

# Funções de Cálculo
def calcular_dias_para_vencer(df):
    AQ1 = datetime.now().date()
    df["VALIDADE"] = pd.to_datetime(df["VALIDADE"]).dt.date
    df["DIAS PARA VENCER"] = df["VALIDADE"].apply(lambda x: (x - AQ1).days if not pd.isna(x) and (x - AQ1).days > 0 else 0)
    return df

def calcular_medias_e_projeções(df):
    df["VENDA MEDIA MENSAL"] = df["VENDA ULTIMOS 12 MESES (QTDE)"] / 12
    df["MAIOR VENDA"] = df[["VENDA ULTIMOS 30 DIAS (QTDE)", "VENDA MEDIA MENSAL"]].max(axis=1)
    df["VENDA DIÁRIA"] = df["MAIOR VENDA"] / 30
    df["PROJEÇÃO DE VENDA"] = df["DIAS PARA VENCER"] * df["VENDA DIÁRIA"]
    df["EST > PROJ"] = df.apply(lambda row: 1 if row["ESTOQUE ATUAL"] > row["PROJEÇÃO DE VENDA"] else 0, axis=1)
    return df

def calcular_estoque_excedente_e_custos(df):
    df["ESTOQUE EXCEDENTE"] = df.apply(lambda row: 0 if row["EST > PROJ"] == 0 else row["ESTOQUE ATUAL"] - row["PROJEÇÃO DE VENDA"], axis=1)
    df["CUSTO TOTAL EXCEDENTE"] = df["ESTOQUE EXCEDENTE"] * df["PREÇO DE CUSTO"]
    df["VENDA TOTAL"] = df["PREÇO DE VENDA"] * df["ESTOQUE EXCEDENTE"]
    return df

def adicionar_colunas_comparativas(df):
    df["60 DIAS < PRAZO"] = df["DIAS PARA VENCER"].apply(lambda x: "SIM" if x <= 60 else "NÃO")
    df["EST ATUAL < X EMB"] = df.apply(lambda row: "SIM" if row["ESTOQUE ATUAL"] < (1 * row["QTDE NA EMBALAGEM"]) else "NÃO", axis=1)
    return df

# Aplicando as funções de cálculo
df_relatorio_calculos = df_relatorio_renomeado.copy()
df_relatorio_calculos = calcular_dias_para_vencer(df_relatorio_calculos)
df_relatorio_calculos = calcular_medias_e_projeções(df_relatorio_calculos)
df_relatorio_calculos = df_relatorio_calculos.merge(df_preços_renomeado, on=["CODIGO", "LOJA"], how="left")
df_relatorio_calculos = calcular_estoque_excedente_e_custos(df_relatorio_calculos)
df_relatorio_calculos = adicionar_colunas_comparativas(df_relatorio_calculos)

# Funções de Agregação e Formatação
def agregar_dados_por_loja(df):
    agregados = {
        "EST > PROJ": lambda x: (x == 1).sum(),
        "ESTOQUE EXCEDENTE": "sum",  
        "CUSTO TOTAL EXCEDENTE": "sum",        
        "VENDA TOTAL": "sum"        
    }
    df_agregado = df.groupby("LOJA").agg(agregados).reset_index()
    df_agregado.columns = ["LOJA", "QTDE ITENS", "QTDE EXCEDENTE", "VALOR CUSTO TOTAL", "VALOR VENDA TOTAL"]
    return df_agregado

def adicionar_linha_total(df):
    total_itens = df["QTDE ITENS"].sum()
    total_excedente = df["QTDE EXCEDENTE"].sum()
    total_custo = df["VALOR CUSTO TOTAL"].sum()
    total_venda = df["VALOR VENDA TOTAL"].sum()
    df.loc["Total"] = ["Total:", total_itens, total_excedente, total_custo, total_venda]
    return df

def formatar_qtde_excedente(df):
    df["QTDE EXCEDENTE"] = pd.to_numeric(df["QTDE EXCEDENTE"], errors="coerce").fillna(0).astype(int)
    return df

# Aplicando as funções de agregação e formatação
df_relatorio_resumo = df_relatorio_calculos.copy()
df_relatorio_resumo = agregar_dados_por_loja(df_relatorio_resumo)
df_relatorio_resumo = adicionar_linha_total(df_relatorio_resumo)
df_relatorio_resumo = formatar_qtde_excedente(df_relatorio_resumo)

# Funções de Processamento e Análise
def processar_dados_por_loja(df, loja):
    def formatar_numero(num):
        if num == 0:
            return 0
        elif num < 1:
            return round(num, 2)
        else:
            return round(num, 1)

    df['ESTOQUE ATUAL'] = df['ESTOQUE ATUAL'].apply(formatar_numero)
    df['ESTOQUE EXCEDENTE'] = df['ESTOQUE EXCEDENTE'].apply(formatar_numero)

    df['VALIDADE'] = pd.to_datetime(df['VALIDADE'], dayfirst=True)
    df['VALIDADE'] = df['VALIDADE'].dt.strftime('%d/%m/%Y')

    df_filtrado = df[((df["LOJA"] == loja) & (df["CUSTO TOTAL EXCEDENTE"] > 0))].copy()
    df_filtrado['ÚLTIMA AÇÃO'] = ''  # Adicionando a coluna "ÚLTIMA AÇÃO" vazia
    df_filtrado['NOVA AÇÃO'] = ''  # Adicionando a coluna "NOVA AÇÃO" vazia

    df_ordenado = df_filtrado[["CODIGO", "NOME", "SETOR", "VALIDADE", "ESTOQUE ATUAL", "ESTOQUE EXCEDENTE", "CUSTO TOTAL EXCEDENTE", "ÚLTIMA AÇÃO", "NOVA AÇÃO"]].sort_values(by=["CUSTO TOTAL EXCEDENTE"], ascending=False)
    return df_ordenado

# Processando o resumo de itens com estoque excedente por loja
resultado_loja_SMJ = processar_dados_por_loja(df_relatorio_calculos, "TRESMANN - SMJ")
resultado_loja_STT = processar_dados_por_loja(df_relatorio_calculos, "TRESMANN - STT")
resultado_loja_VIX = processar_dados_por_loja(df_relatorio_calculos, "TRESMANN - VIX")

# Funções de Salvamento e Formatação de Arquivo Excel
def ajustar_formato_colunas(escritor, nome_aba, dataframe, colunas_moeda):
    """Ajusta a largura e o formato das colunas."""
    for coluna in dataframe.columns:
        # Aplicar 'map' na série (coluna) e calcular a largura máxima
        largura_coluna = max(dataframe[coluna].astype(str).map(len).max(), len(coluna)) + 2
        indice_coluna = dataframe.columns.get_loc(coluna) + 1
        escritor.sheets[nome_aba].column_dimensions[get_column_letter(indice_coluna)].width = largura_coluna

        # Seus códigos para formatar as colunas como moeda ou milhar devem ser inseridos aqui
        if coluna in colunas_moeda:
            aplicar_formato_moeda(escritor, nome_aba, indice_coluna, dataframe.shape[0])

def aplicar_formato_moeda(escritor, nome_aba, indice_coluna, quantidade_linhas):
    """Aplica o formato de moeda às células."""
    for linha in range(2, quantidade_linhas + 2):
        celula = escritor.sheets[nome_aba].cell(row=linha, column=indice_coluna)
        celula.number_format = "R$#,##0.00"

def salvar_dfs_no_excel(escritor, informacoes_dfs):
    """Salva os DataFrames em suas respectivas abas e ajusta as colunas."""
    for info_df in informacoes_dfs:
        info_df['dataframe'].to_excel(escritor, sheet_name=info_df['nome_aba'], index=False)
        ajustar_formato_colunas(escritor, info_df['nome_aba'], info_df['dataframe'], colunas_moeda)
    
    print("Relatório de Estoque Gerado!")

def atualizar_historico_resumo():
    """Atualiza a aba 'HISTÓRICO RESUMO' com as primeiras três linhas de dados do 'RESUMO'."""
    df_resumo_relatorio = pd.read_excel(CAMINHO_ARQUIVO, engine='openpyxl', sheet_name='RESUMO')

    try:
        df_historico_existente = pd.read_excel(CAMINHO_ARQUIVO, sheet_name='HISTÓRICO RESUMO')
    except (FileNotFoundError, ValueError):  # Trata tanto a ausência do arquivo quanto a da aba
        df_historico_existente = pd.DataFrame()

    # Pegando apenas as 3 primeiras linhas de 'RESUMO' e criando uma cópia independente
    df_resumo_atualizado = df_resumo_relatorio.head(3).copy()
    df_resumo_atualizado['DATA'] = datetime.now().strftime('%d/%m/%Y')
    df_historico_atualizado = pd.concat([df_historico_existente, df_resumo_atualizado], ignore_index=True)

    return df_historico_atualizado

def salvar_historico_acoes():
    try:
        df_historico_ações = pd.read_excel(CAMINHO_ARQUIVO, engine='openpyxl', sheet_name='HISTÓRICO AÇÕES')
    except (FileNotFoundError, ValueError):  # Trata tanto a ausência do arquivo quanto a da aba
        df_historico_ações = pd.DataFrame()

    sheets_lojas = ['TRESMANN - SMJ', 'TRESMANN - STT', 'TRESMANN - VIX']

    for sheet in sheets_lojas:
        df = pd.read_excel(CAMINHO_ARQUIVO, sheet_name=sheet)
        # Filtrar as linhas onde "NOVA AÇÃO" não está vazia
        df_filtrado = df[df['NOVA AÇÃO'].notna() & (df['NOVA AÇÃO'] != '')].copy()
        if not df_filtrado.empty:
            df_filtrado['LOJA'] = sheet  # Nome da sheet como valor da coluna "LOJA"
            df_filtrado['DATA'] = datetime.now().strftime('%d/%m/%Y')  # Adicionando a data atual
            df_filtrado = df_filtrado[['CODIGO', 'LOJA', 'DATA', 'NOVA AÇÃO']]
            df_historico_ações = pd.concat([df_historico_ações, df_filtrado], ignore_index=True)

    return df_historico_ações

def verificar_existencia_arquivo():
    """Verifica se o arquivo Excel existe."""
    try:
        # Tenta ler o arquivo
        pd.read_excel(CAMINHO_ARQUIVO, engine='openpyxl')
        return True
    except FileNotFoundError:
        # Se o arquivo não existir, retorna False
        return False
    
informacoes_dfs = [
    {"nome_aba": 'RESUMO', "dataframe": df_relatorio_resumo},
    {"nome_aba": 'TRESMANN - SMJ', "dataframe": resultado_loja_SMJ},
    {"nome_aba": 'TRESMANN - STT', "dataframe": resultado_loja_STT},
    {"nome_aba": 'TRESMANN - VIX', "dataframe": resultado_loja_VIX}
]

# Código principal para salvar os DataFrames
if verificar_existencia_arquivo():
    df_historico_atualizado = atualizar_historico_resumo()
    df_historico_acoes = salvar_historico_acoes()

    with pd.ExcelWriter(CAMINHO_ARQUIVO, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Primeiro salvar todas as abas, exceto 'HISTÓRICO RESUMO'
        salvar_dfs_no_excel(writer, informacoes_dfs)

        # Depois salvar o 'HISTÓRICO RESUMO' atualizado
        df_historico_atualizado.to_excel(writer, sheet_name='HISTÓRICO RESUMO', index=False)
        ajustar_formato_colunas(writer, 'HISTÓRICO RESUMO', df_historico_atualizado, colunas_moeda)

        if not df_historico_acoes.empty:
            df_historico_acoes.to_excel(writer, sheet_name='HISTÓRICO AÇÕES', index=False)
            ajustar_formato_colunas(writer, 'HISTÓRICO AÇÕES', df_historico_acoes, colunas_moeda)
else:
    with pd.ExcelWriter(CAMINHO_ARQUIVO, engine='openpyxl') as writer:
        # Se o arquivo não existir, apenas salvar os DataFrames
        salvar_dfs_no_excel(writer, informacoes_dfs)

enviar_email()