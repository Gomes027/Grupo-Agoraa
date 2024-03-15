import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font

class AjusteDeColuna:
    @staticmethod
    def ajustar_largura_colunas(sheet, dataframe, buffer=4):
        """
        Ajusta a largura das colunas de uma folha Excel com base no comprimento dos dados
        e dos cabeçalhos das colunas, usando openpyxl.

        Parâmetros:
            sheet (openpyxl worksheet): A folha de cálculo para ajustar as larguras das colunas.
            dataframe (pd.DataFrame): O DataFrame para basear os ajustes das larguras das colunas.
            buffer (int): Um buffer adicional para adicionar ao comprimento máximo para evitar corte.
        """
        for column in dataframe.columns:
            max_length = max(dataframe[column].astype(str).map(len).max(), len(column)) + buffer
            column_index = dataframe.columns.get_loc(column) + 1
            col_letter = get_column_letter(column_index)
            sheet.column_dimensions[col_letter].width = max_length

class GerenciadorDeCotacoes:
    def __init__(self, pasta_fornecedores, diretorio_original, diretorio_destino):
        self.pasta_fornecedores = pasta_fornecedores
        self.diretorio_original = diretorio_original
        self.diretorio_destino = diretorio_destino
        self.relatorio_cotacao = None

    def remover_arquivo_existente(self):
        try:
            os.remove(self.diretorio_destino)
        except FileNotFoundError:
            pass

    def carregar_relatorio_original(self):
        self.relatorio_cotacao = pd.read_excel(self.diretorio_original)

    def processar_cotacoes_fornecedores(self):
        for arquivo in os.listdir(self.pasta_fornecedores):
            if arquivo.endswith('.xlsx'):
                self.processar_arquivo_fornecedor(arquivo)

    def processar_arquivo_fornecedor(self, arquivo):
        nome_coluna, _ = os.path.splitext(arquivo)
        nome_coluna = nome_coluna.split("-")[0].replace(" ", "")

        caminho_arquivo = os.path.join(self.pasta_fornecedores, arquivo)
        fornecedor_cotacao = pd.read_excel(caminho_arquivo, header=2)

        fornecedor_cotacao.rename(columns={fornecedor_cotacao.columns[3]: nome_coluna}, inplace=True)
        fornecedor_cotacao.rename(columns={'COD.': 'COD', 'COD BARRAS': 'COD. BARRA.'}, inplace=True)
        
        self.relatorio_cotacao = self.relatorio_cotacao.merge(fornecedor_cotacao, how="left")
        print(f"{nome_coluna} adicionado ao comparativo de cotações.")

    def adicionar_colunas_calculadas(self):
        def validacao(value):
            return "" if pd.isna(value) or value == "" else "SIM"

        def custom_function(columns):
            try:
                return np.std(columns, ddof=0)
            except:
                return ""
        
        colunas_para_calcular = self.relatorio_cotacao.columns[self.relatorio_cotacao.columns.get_loc('MENOR') + 1:]
        self.relatorio_cotacao[colunas_para_calcular] = self.relatorio_cotacao[colunas_para_calcular].apply(pd.to_numeric, errors='coerce')
        col_range = self.relatorio_cotacao.loc[:, colunas_para_calcular]

        self.relatorio_cotacao['MENOR'] = self.relatorio_cotacao[colunas_para_calcular].min(axis=1)
        self.relatorio_cotacao['QTDE COTAÇÃO'] = self.relatorio_cotacao[colunas_para_calcular].apply(lambda row: row.count() if pd.notna(row).any() else None, axis=1)

        self.relatorio_cotacao['MELHOR FORN,'] = None
        for i, row in self.relatorio_cotacao.iterrows():
            valid_row = row[colunas_para_calcular].dropna()
            if not valid_row.empty:
                # Garantir que estamos lidando com valores numéricos antes de chamar idxmin
                valid_row = valid_row.astype(float)
                self.relatorio_cotacao.at[i, 'MELHOR FORN,'] = valid_row.idxmin()

        self.relatorio_cotacao['VALIDAÇÃO'] = self.relatorio_cotacao['MENOR'].apply(validacao)
        self.relatorio_cotacao['DESV. PADRÃO'] = col_range.apply(custom_function, axis=1)

    def salvar_relatorio(self, nome_arquivo_excel, sheet_name='Comparativo de Cotações'):
        # Salvando o DataFrame em um arquivo Excel
        self.relatorio_cotacao.to_excel(nome_arquivo_excel, sheet_name=sheet_name, index=False)
        
        # Carregando o arquivo Excel para ajustar formatações com openpyxl
        workbook = load_workbook(nome_arquivo_excel)
        sheet = workbook[sheet_name]
        
        # Definições de estilo
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="BDBDBD", end_color="BDBDBD", fill_type="solid")
        row_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        row_fill_2 = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
        
        # Aplicando formatação ao cabeçalho
        for col in range(1, len(self.relatorio_cotacao.columns) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        # Aplicando formatação às linhas
        for row in range(2, sheet.max_row + 1):
            for col in range(1, len(self.relatorio_cotacao.columns) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.fill = row_fill_1 if row % 2 == 0 else row_fill_2
                cell.border = thin_border

        # Aplicar formato numérico
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = numbers.FORMAT_NUMBER
        
        # Ajustando largura das colunas
        AjusteDeColuna.ajustar_largura_colunas(sheet, self.relatorio_cotacao)

        # Salvando as mudanças
        workbook.save(nome_arquivo_excel)

    def executar(self):
        self.remover_arquivo_existente()
        self.carregar_relatorio_original()
        self.processar_cotacoes_fornecedores()
        self.adicionar_colunas_calculadas()
        self.salvar_relatorio(self.diretorio_destino)
        try:
            os.system(f'start excel "{self.diretorio_destino}"')
        except Exception as e:
            print(f"Erro ao abrir o Excel: {e}")

gerenciador = GerenciadorDeCotacoes(r'Fornecedores', r'Comparativo\Original\COMPARATIVO DE COTAÇÕES - ORIGINAL v2.xlsx', r'Comparativo\Original\COMPARATIVO DE COTAÇÕES - v2.xlsx')
gerenciador.executar()