import os
import json
import pytz
import socket
import sqlite3
import pandas as pd
from pathlib import Path
from datetime import datetime
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, NamedStyle

class BancoDeDados:
    def __init__(self, db_path=r'DataBase\registros_cadastro.db'):
        self.db_path = db_path
        self.criar_banco()

    def buscar_dados(self):
        """
        Este método busca todos os registros na tabela especificada e retorna uma lista de tuplas com os dados.
        """
        conexao = sqlite3.connect(self.db_path)
        cursor = conexao.cursor()
        
        # Exemplo de consulta que busca dados específicos. Adapte conforme necessário.
        cursor.execute("SELECT ID, MOTIVO_CADASTRO, EAN, DESCRICAO_COMPLETA FROM cadastro_definitivo")
        
        dados = cursor.fetchall()
        
        conexao.close()
        return dados

    def criar_banco(self):
        conexao = sqlite3.connect(self.db_path)
        cursor = conexao.cursor()
        # Cria tabela definitiva
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS cadastro_definitivo (
            ID INTEGER PRIMARY KEY,
            DATA_HORA TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIORIDADE TEXT,           
            MOTIVO_CADASTRO TEXT,
            EAN TEXT,
            DESCRICAO_COMPLETA TEXT,
            DESCRICAO_CONSUMIDOR TEXT,
            FORNECEDOR TEXT,
            MARCA TEXT,
            GRAMATURA TEXT,
            EMBALAGEM_COMPRA TEXT,
            NCM TEXT,
            CODIGO_INTERNO TEXT,
            SETOR TEXT,
            GRUPO TEXT,
            SUBGRUPO TEXT,
            CATEGORIA TEXT,        
            "MIX SMJ" TEXT,
            "MIX STT" TEXT,
            "MIX VIX" TEXT,
            "MIX MCP" TEXT,
            "TR SMJ" TEXT,
            "TR STT" TEXT,
            "TR VIX" TEXT,
            "TR MCP" TEXT,
            "AP" TEXT
        )
        ''')
        # Cria tabela temporária
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS cadastro_temporario (
            ID INTEGER PRIMARY KEY,
            DATA_HORA TIMESTAMP DEFAULT CURRENT_TIMESTAMP,  
            PRIORIDADE TEXT,         
            MOTIVO_CADASTRO TEXT,
            EAN TEXT,
            DESCRICAO_COMPLETA TEXT,
            DESCRICAO_CONSUMIDOR TEXT,
            FORNECEDOR TEXT,
            MARCA TEXT,
            GRAMATURA TEXT,
            EMBALAGEM_COMPRA TEXT,
            NCM TEXT,
            CODIGO_INTERNO TEXT,
            SETOR TEXT,
            GRUPO TEXT,
            SUBGRUPO TEXT,
            CATEGORIA TEXT,         
            "MIX SMJ" TEXT,
            "MIX STT" TEXT,
            "MIX VIX" TEXT,
            "MIX MCP" TEXT,
            "TR SMJ" TEXT,
            "TR STT" TEXT,
            "TR VIX" TEXT,
            "TR MCP" TEXT,
            "AP" TEXT
        )
        ''')
        conexao.commit()
        conexao.close()

    def inserir_dados_temp(self, dados):
        try:
            self.inserir_dados(dados, "cadastro_temporario")
            return True
        except Exception as e:
            print(f"Erro ao inserir dados: {e}")
            return False

    def inserir_dados_definitivo(self, dados):
        self.inserir_dados(dados, "cadastro_definitivo")

    def inserir_dados(self, dados, tabela):
        conexao = sqlite3.connect(self.db_path)
        cursor = conexao.cursor()

        # Obter a data e hora atual no fuso horário de Brasília
        fuso_horario = pytz.timezone('America/Sao_Paulo')
        data_hora_br = datetime.now(fuso_horario).strftime('%Y-%m-%d %H:%M:%S')

        # Adiciona a data_hora aos dados, assumindo que não é uma chave nos dados originais
        dados['DATA_HORA'] = data_hora_br

        colunas = ', '.join([f'"{coluna}"' for coluna in dados.keys()])
        placeholders = ', '.join(['?'] * len(dados))
        valores = list(dados.values())

        cursor.execute(f"INSERT INTO {tabela} ({colunas}) VALUES ({placeholders})", valores)

        conexao.commit()
        conexao.close()

    def exportar_arquivo(self, nome_arquivo):
        # Verificar se há dados na tabela temporária
        conexao = sqlite3.connect(self.db_path)
        cursor = conexao.cursor()
        cursor.execute("SELECT COUNT(*) FROM cadastro_temporario")
        numero_de_linhas = cursor.fetchone()[0]
        conexao.close()
        
        if numero_de_linhas == 0:
            return 0
        else:
            try:
                self.mover_dados_para_definitivo()
                self.exportar_para_excel(nome_arquivo)
                self.exportar_dados()
                return True
            except Exception as e:
                print(f"Erro ao exportar dados: {e}")
                return False

    def mover_dados_para_definitivo(self):
        conexao = sqlite3.connect(self.db_path)
        cursor = conexao.cursor()
        cursor.execute("SELECT * FROM cadastro_temporario")
        dados_temp = cursor.fetchall()

        colunas_sql = ", ".join(["DATA_HORA", "PRIORIDADE", "MOTIVO_CADASTRO", "EAN", "DESCRICAO_COMPLETA", "DESCRICAO_CONSUMIDOR", "FORNECEDOR", "MARCA", 
                "GRAMATURA", "EMBALAGEM_COMPRA", "NCM", "CODIGO_INTERNO", "SETOR", "GRUPO", "SUBGRUPO", "CATEGORIA",
                "\"MIX SMJ\"", "\"MIX STT\"", "\"MIX VIX\"", "\"MIX MCP\"", "\"TR SMJ\"", "\"TR STT\"", "\"TR VIX\"", "\"TR MCP\"", "\"AP\""])
        placeholders = ", ".join(["?"] * 25)  # Ajuste no número de placeholders conforme necessário

        for dado in dados_temp:
            valores = dado[1:]  # Exclui o ID primário
            cursor.execute(f"INSERT INTO cadastro_definitivo ({colunas_sql}) VALUES ({placeholders})", valores)

        conexao.commit()
        conexao.close()

    def limpar_tabela_temp(self):
        conexao = sqlite3.connect(self.db_path)
        cursor = conexao.cursor()
        cursor.execute("DELETE FROM cadastro_temporario")
        conexao.commit()
        conexao.close()

    def ajustar_formato_colunas(self, workbook, nome_aba, dataframe):
        aba = workbook[nome_aba]
        for coluna in dataframe.columns:
            largura_coluna = max(
                dataframe[coluna].astype(str).apply(len).max(),
                len(coluna)
            ) + 8
            indice_coluna = dataframe.columns.get_loc(coluna) + 1
            aba.column_dimensions[get_column_letter(indice_coluna)].width = largura_coluna

    def exportar_para_excel(self, nome_arquivo):
        pasta_downloads = str(Path.home() / "Downloads")
        nome_arquivo_excel = os.path.join(pasta_downloads, nome_arquivo)

        conexao = sqlite3.connect(self.db_path)
        try:
            df = pd.read_sql_query("SELECT * FROM cadastro_temporario", conexao)

            colunas_numericas = ['EAN', 'EMBALAGEM_COMPRA', 'NCM']
            for coluna in colunas_numericas:
                df[coluna] = pd.to_numeric(df[coluna], errors='coerce')

            # Exclui a coluna DATA_HORA
            df = df.drop(columns=['DATA_HORA', 'ID'])

            df.columns = df.columns.str.upper()

            df.to_excel(nome_arquivo_excel, index=False, engine='openpyxl', sheet_name='Dados Cadastro')

            workbook = load_workbook(nome_arquivo_excel)
            sheet = workbook.active

            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

            header_fill = PatternFill(start_color="BDBDBD", end_color="BDBDBD", fill_type="solid")
            row_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            row_fill_2 = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")

            for col in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.border = thin_border

            for row in range(2, sheet.max_row + 1):
                for col in range(1, len(df.columns) + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.fill = row_fill_1 if row % 2 == 0 else row_fill_2
                    cell.border = thin_border

            self.ajustar_formato_colunas(workbook, 'Dados Cadastro', df)

            workbook.save(nome_arquivo_excel)
            print(f"Dados exportados com sucesso para {nome_arquivo_excel}")
        except Exception as e:
            print(f"Erro ao exportar dados: {e}")
        finally:
            conexao.close()

    def send_data_to_server(self, data):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as client_socket:
                client_socket.settimeout(5)  # Define um timeout de 5 segundos
                client_socket.connect(('127.0.0.1', 65432))
                request_data = {"action": "insert", "data": data}
                client_socket.sendall(json.dumps(request_data).encode('utf-8'))
                
                # Aguarda uma resposta do servidor
                response = client_socket.recv(1024).decode('utf-8')
                response_data = json.loads(response)
                
                if response_data["status"] == "success":
                    print("Dados inseridos com sucesso!")
                    return True
                else:
                    print("Falha ao inserir dados:", response_data["message"])
                    return False
        except socket.timeout:
            print("Tempo limite excedido ao aguardar resposta do servidor.")
            return False
        except Exception as e:
            print("Erro ao conectar ao servidor:", e)
            return False

    def exportar_dados(self):
        # Suponho que as colunas selecionadas incluam "MIX" e "TR" para cada loja e "AP" no final.
        colunas_selecionadas = (
            "PRIORIDADE", "EAN", "DESCRICAO_COMPLETA", "DESCRICAO_CONSUMIDOR",
            "FORNECEDOR", "MARCA", "GRAMATURA", "EMBALAGEM_COMPRA", "NCM",
            "SETOR", "GRUPO", "SUBGRUPO", "CATEGORIA",
            "\"MIX SMJ\"", "\"MIX STT\"", "\"MIX VIX\"", "\"MIX MCP\"",
            "\"TR SMJ\"", "\"TR STT\"", "\"TR VIX\"", "\"TR MCP\"", "\"AP\""
        )

        conexao = sqlite3.connect(self.db_path)
        try:
            cursor = conexao.cursor()
            query = f"SELECT {', '.join(colunas_selecionadas)} FROM cadastro_temporario"
            cursor.execute(query)
            dados = cursor.fetchall()

            todos_enviados_com_sucesso = True

            for dado in dados:
                # Descompactar os valores, exceto os últimos oito que são relativos a MIX, TR e AP
                *informacoes_basicas, mix_smj, mix_stt, mix_vix, mix_mcp, tr_smj, tr_stt, tr_vix, tr_mcp, ap = dado

                # Processar MIX para coletar as siglas das lojas onde é "sim"
                lojas_mix = [loja for loja, mix in zip(["SMJ", "STT", "VIX", "MCP"], [mix_smj, mix_stt, mix_vix, mix_mcp]) if mix == "SIM"]
                mix = ', '.join(lojas_mix)  # Junta as siglas com vírgula

                # Processar TR para coletar as siglas das lojas onde é "sim"
                lojas_tr = [loja for loja, tr in zip(["SMJ", "STT", "VIX", "MCP"], [tr_smj, tr_stt, tr_vix, tr_mcp]) if tr == "SIM"]
                tr = ', '.join(lojas_tr)  # Junta as siglas com vírgula

                # Preparar o dado a ser enviado (convertendo tudo para lista antes da concatenação)
                dado_para_enviar = list(informacoes_basicas) + [mix, tr, ap]

                sucesso = self.send_data_to_server(dado_para_enviar)
                if not sucesso:
                    todos_enviados_com_sucesso = False
                    break  # Interrompe o envio dos dados restantes

            if todos_enviados_com_sucesso:
                messagebox.showinfo("Sucesso", "Dados exportados com sucesso para o servidor.")
                self.limpar_tabela_temp()
            else:
                messagebox.showerror("Erro", "Servidor não conectado. Tente novamente mais tarde.")
                self.limpar_tabela_temp()

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar dados para o servidor: {e}")
        finally:
            conexao.close()

    def verificar_ean_existente(self, ean):
        conexao = sqlite3.connect('DataBase/dados_internos.db')
        cursor = conexao.cursor()
        
        cursor.execute("SELECT CODIGO, NOME FROM codigos_internos WHERE \"CODG DE BARRAS\" = ?", (ean,))
        resultado = cursor.fetchone()
        
        conexao.close()
        
        if resultado:
            codigo, nome = resultado
            return True, codigo, nome
        else:
            return False, None, None